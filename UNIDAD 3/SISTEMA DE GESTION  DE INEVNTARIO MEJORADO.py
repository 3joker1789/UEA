#Tarea: Sistema de Gestión de Inventarios Mejorado

#Objetivo: Mejorar el sistema de gestión de inventarios desarrollado anteriormente para que utilice archivos para almacenar y recuperar la información del inventario y maneje excepciones durante la lectura y escritura de archivos.

#Nuevos Requisitos:

    #Almacenamiento de Inventarios en Archivos:
        #Modificar la clase Inventario para que al añadir,actualizar, o eliminar productos, estas modificaciones se reflejen en un archivo de texto (por ejemplo, inventario.txt).

    #Recuperación de Inventarios desde Archivos:
        #Al iniciar el programa, cargar automáticamente los productos existentes en inventario.txt para reconstruir el inventario.

    #Manejo de Excepciones:
        #Implementar manejo de excepciones para capturar y tratar adecuadamente posibles errores durante la manipulación de archivos, como FileNotFoundError y PermissionError.
        #Asegurar que el programa maneje casos en los que el archivo de inventario no exista, creándolo si es necesario.

    #Modificaciones a la Interfaz de Usuario en la Consola:
        #Actualizar la interfaz de usuario para notificar al usuario sobre el éxito o fallo de operaciones de archivo (por ejemplo, notificar al usuario cuando un producto se añade exitosamente al archivo de inventario).

#Instrucciones Adicionales:

    #Mantén la organización y claridad del código, asegurando que todas las modificaciones estén bien comentadas para explicar el funcionamiento del manejo de archivos y excepciones.
    #Realiza pruebas exhaustivas para asegurarte de que el programa puede manejar situaciones como archivos corruptos, falta de permisos de escritura, y más.

import os
from openpyxl import Workbook, load_workbook

class Inventario:
    def __init__(self, archivo='inventario.xlsx'):
        self.archivo = archivo
        self.productos = {}
        self.cargar_inventario()

    def cargar_inventario(self):
        """Carga los productos desde un archivo al iniciar el programa."""
        if not os.path.exists(self.archivo):
            self.crear_archivo()

        if self.archivo.endswith('.txt'):
            self.cargar_desde_txt()
        elif self.archivo.endswith('.xlsx'):
            self.cargar_desde_excel()
        else:
            print("Formato de archivo no soportado.")

    def cargar_desde_txt(self):
        """Carga los productos desde un archivo de texto."""
        try:
            with open(self.archivo, 'r') as file:
                for linea in file:
                    id_producto, nombre, cantidad, precio = linea.strip().split(',')
                    self.productos[id_producto] = {
                        'nombre': nombre,
                        'cantidad': int(cantidad),
                        'precio': float(precio)
                    }
        except Exception as e:
            print(f"Error al cargar el inventario desde texto: {e}")

    def cargar_desde_excel(self):
        """Carga los productos desde un archivo de Excel usando openpyxl."""
        try:
            wb = load_workbook(self.archivo)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                id_producto, nombre, cantidad, precio = row
                self.productos[str(id_producto)] = {
                    'nombre': nombre,
                    'cantidad': int(cantidad),
                    'precio': float(precio)
                }
        except Exception as e:
            print(f"Error al cargar el inventario desde Excel: {e}")

    def guardar_inventario(self):
        """Guarda los productos en el archivo."""
        if self.archivo.endswith('.txt'):
            self.guardar_en_txt()
        elif self.archivo.endswith('.xlsx'):
            self.guardar_en_excel()
        else:
            print("Formato de archivo no soportado.")

    def guardar_en_txt(self):
        """Guarda los productos en un archivo de texto."""
        try:
            with open(self.archivo, 'w') as file:
                for id_producto, datos in self.productos.items():
                    file.write(f"{id_producto},{datos['nombre']},{datos['cantidad']},{datos['precio']}\n")
            print("Inventario guardado exitosamente en texto.")
        except PermissionError:
            print("Error: Permiso denegado para escribir en el archivo.")
        except Exception as e:
            print(f"Error al guardar el inventario en texto: {e}")

    def guardar_en_excel(self):
        """Guarda los productos en un archivo de Excel usando openpyxl."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(['ID', 'Nombre', 'Cantidad', 'Precio'])  # Encabezados
            for id_producto, datos in self.productos.items():
                ws.append([id_producto, datos['nombre'], datos['cantidad'], datos['precio']])
            wb.save(self.archivo)
            print("Inventario guardado exitosamente en Excel.")
        except PermissionError:
            print("Error: Permiso denegado para escribir en el archivo.")
        except Exception as e:
            print(f"Error al guardar el inventario en Excel: {e}")

    def crear_archivo(self):
        """Crea un archivo vacío si no existe."""
        if self.archivo.endswith('.txt'):
            self.crear_archivo_txt()
        elif self.archivo.endswith('.xlsx'):
            self.crear_archivo_excel()
        else:
            print("Formato de archivo no soportado.")

    def crear_archivo_txt(self):
        """Crea un archivo de texto vacío si no existe."""
        try:
            with open(self.archivo, 'w') as file:
                pass
        except PermissionError:
            print("Error: Permiso denegado para crear el archivo.")
        except Exception as e:
            print(f"Error al crear el archivo de texto: {e}")

    def crear_archivo_excel(self):
        """Crea un archivo de Excel vacío si no existe."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(['ID', 'Nombre', 'Cantidad', 'Precio'])  # Encabezados
            wb.save(self.archivo)
        except PermissionError:
            print("Error: Permiso denegado para crear el archivo.")
        except Exception as e:
            print(f"Error al crear el archivo de Excel: {e}")

    def añadir_producto(self, id_producto, nombre, cantidad, precio):
        """Añade un nuevo producto al inventario y guarda los cambios en el archivo."""
        if id_producto in self.productos:
            print("El producto ya existe en el inventario.")
        else:
            self.productos[id_producto] = {
                'nombre': nombre,
                'cantidad': cantidad,
                'precio': precio
            }
            self.guardar_inventario()

    def actualizar_producto(self, id_producto, cantidad, precio):
        """Actualiza la cantidad y el precio de un producto existente y guarda los cambios en el archivo."""
        if id_producto in self.productos:
            self.productos[id_producto]['cantidad'] = cantidad
            self.productos[id_producto]['precio'] = precio
            self.guardar_inventario()
        else:
            print("El producto no existe en el inventario.")

    def eliminar_producto(self, id_producto):
        """Elimina un producto del inventario y guarda los cambios en el archivo."""
        if id_producto in self.productos:
            del self.productos[id_producto]
            self.guardar_inventario()
        else:
            print("El producto no existe en el inventario.")

# Ejemplo de uso:
inventario = Inventario(archivo='inventario.xlsx')

# Añadir un producto
inventario.añadir_producto('1', 'Compresor LG QK-134', 15, 150.00)

# Actualizar un producto
inventario.actualizar_producto('001', 8, 949.99)

# Eliminar un producto
inventario.eliminar_producto('001')


